# -*- coding: utf-8 -*-
"""
Gerador do Excel de Nomenclatura AOB 2026 (PT) a partir do nomINT2026.txt (EN).

Regras aplicadas (validadas com o utilizador):
- Colunas: Classe | Seccao AOB | Codigo BVA | Mutacao
- Individuais: numero par em Seccao AOB (2, 4, 6, ...)
- Equipas:    numero impar em Seccao AOB (1, 3, 5, ...) -- contador proprio
- NYA (Not Yet Accepted) omitidas do output
- DM jade -> jade DF
- "aves da serie violeta" removida das equipas Eyering
- Uniformizado "D factor violeta" -> "D violeta"
- Layout em blocos por especie (como 2024)
- Grupo Juizes mantido (especificidade AOB)
"""

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).parent
TXT = BASE / "nomINT2026.txt"
OUT = BASE / "Classes_Agapornis_2026_PT.xlsx"


# ============================================================
# GLOSSARIO DE TRADUCAO
# ============================================================

# Mutacoes primarias e modificadores (mantem-se em ingles, excepto onde traduzido)
# A ordem canonica na celula PT:
#   A. Especie <cor-base> [D|DD] <face laranja|pale headed|slaty|violeta>
#     <opalino> <mut-primaria> [misty DF]

COR_BASE_PT = {
    # cores base "puras"
    "green": "verde",
    "aqua": "aqua",
    "turquoise": "turquesa",
    "teal": "teal",
    # blue: depende de especie -> tratamento especial
}

# Modificadores adjectivos (aparecem antes/depois consoante o padrao)
MODIFICADORES_PT = {
    "orange face": "face laranja",
    "pale headed": "pale headed",
    "slaty": "slaty",
    "violet": "violeta",
    "opaline": "opalino",
}

# Mutacoes principais (uma por linha logica)
MUTACOES_PRIMARIAS_PT = {
    "marbled": "marbled",
    "dilute": "diluido",
    "bronze fallow": "bronze fallow",
    "pale fallow": "pale fallow",
    "dun fallow": "dun fallow",
    "cinnamon": "canela",
    "pallid": "pálido",
    "pale": "pale",
    "pastel": "pastel",
    "dominant pied": "arlequim dominante",
    "recessive pied": "arlequim recessivo",
    "DM jade": "jade DF",
    "crested": "crested",
    "dec": "dec",
    # SL ino / NSL ino (mantidos)
    "SL ino": "SL ino",
    "NSL ino": "NSL ino",
    "ino": "ino",
    # combinacoes com traco
    "opaline-ino": "opalino-ino",
    "cinnamon-ino": "canela-ino",
    "opaline-cinnamon": "opalino-canela",
    "opaline-pallid": "opalino-pálido",
    "opaline-pale": "opalino-pale",
    # edged / euwing / greywing (SF/DF)
    "SF edged": "edged SF",
    "DF edged": "edged DF",
    "SF euwing": "euwing SF",
    "DF euwing": "euwing DF",
    "SL SF greywing": "SL greywing SF",
    "SL DF greywing": "SL greywing DF",
    # misty (sempre DF na aceitacao)
    "DF misty": "misty DF",
}

SEXO_PT = {"male": "macho", "female": "fêmea"}

ESPECIES_TITLE = {
    "roseicollis": "A. Roseicollis",
    "personatus": "A. Personatus",
    "fischeri": "A. Fischeri",
    "nigrigenis": "A. Nigrigenis",
    "lilianae": "A. Lilianae",
    "canus": "A. Canus",
    "taranta": "A. Taranta",
    "pullarius": "A. Pullarius",
}

ESPECIES_LOWER = {k: f"A. {k}" for k in ESPECIES_TITLE}


def blue_pt(especie: str) -> str:
    """Roseicollis: *blue* ; eyering (personatus/fischeri/nigrigenis/lilianae): azul."""
    if especie == "roseicollis":
        return "*blue*"
    return "azul"


def bluebb_pt(especie: str) -> str:
    """Blue1Blue2 -> Azul1Azul2 (so eyering usa)."""
    return "Azul1Azul2"


# ============================================================
# TRADUTOR DE UMA MUTACAO (uma sub-entrada, ex.: "orange face D green")
# ============================================================

def _substitute_word(text: str, en: str, pt: str) -> str:
    """Substitui palavra/expressao inteira. Case-sensitive por defeito;
    para siglas (SF/DF/DM) fica case-insensitive para tolerar typos do PDF."""
    flags = 0
    if re.search(r"\b(SF|DF|DM)\b", en):
        flags = re.IGNORECASE
    return re.sub(rf"(?<!\w){re.escape(en)}(?!\w)", pt, text, flags=flags)


# ============================================================
# ORDEM CANONICA (categorias por token)
# ============================================================
# -1: sexo (macho, fêmea)
#  0: cor-base (verde, *blue*, azul, aqua, turquesa, Azul1Azul2, teal)
#     ou compound base+violeta (jade DF placeholder tambem cai aqui)
#  1: factor (D, DD)
#  2: modificador secundario (face laranja, pale headed, slaty, violeta standalone)
#  3: opalino  (incluido variantes com traco)
#  4: mutacao primaria (marbled, diluido, canela, palido, pale, arlequim*, jade DF,
#                       bronze fallow, pale fallow, dun fallow, edged SF/DF, euwing SF/DF,
#                       SL greywing SF/DF, SL ino, NSL ino, dec, pastel, crested)
#  5: misty DF

TOKEN_CATEGORY = {
    "macho": -1, "fêmea": -1,
    # cor-base simples
    "verde": 0, "*blue*": 0, "azul": 0, "aqua": 0, "turquesa": 0,
    "Azul1Azul2": 0, "teal": 0,
    # compounds atomicos (base+D violeta): sao tokens especiais no sort, depois expandem
    "__BLUE_DVIOLETA__": 0,
    "__AQUA_DVIOLETA__": 0,
    "__TURQUESA_DVIOLETA__": 0,
    "__AZUL1AZUL2_VIOLETA__": 0,
    # factor
    "D": 1, "DD": 1,
    # modificadores secundarios
    "__ORANGE_FACE__": 2, "__PALE_HEADED__": 2,
    "slaty": 2, "__VIOLETA_ONLY__": 2,
    # opalino e variantes com traco (opalino simples)
    "opalino": 3,
    # opalino-<mut> atomicos (mantem-se juntos, mas categoria de opalino)
    "__OP_INO__": 3, "__OP_CIN__": 3, "__OP_PALLID__": 3, "__OP_PALE__": 3,
    # mutacoes primarias
    "marbled": 4, "diluido": 4, "canela": 4, "pálido": 4, "pale": 4,
    "__DOM_PIED__": 4, "__REC_PIED__": 4, "__JADE_DF__": 4,
    "__BRONZE_FALLOW__": 4, "__PALE_FALLOW__": 4, "__DUN_FALLOW__": 4,
    "__SF_EDGED__": 4, "__DF_EDGED__": 4,
    "__SF_EUWING__": 4, "__DF_EUWING__": 4,
    "__SL_SF_GW__": 4, "__SL_DF_GW__": 4,
    "SL": 4,  # SL ino / NSL ino
    "NSL": 4,
    "ino": 4,
    "__CIN_INO__": 4,  # canela-ino
    "dec": 4, "pastel": 4, "crested": 4,
    # misty DF sempre no fim
    "__DF_MISTY__": 5,
}


def _sort_tokens_canonical(tokens: list[str]) -> list[str]:
    """Ordena tokens pela categoria canonica, preservando ordem original dentro
    da mesma categoria (stable sort). Tokens desconhecidos: categoria 4 (mut. primaria)."""
    def cat(tok):
        return TOKEN_CATEGORY.get(tok, 4)
    return sorted(tokens, key=cat)


def traduzir_mutacao(texto_en: str, especie: str) -> str:
    """
    Traduz uma unica mutacao (ex.: "orange face opaline D green")
    para PT canonico (ex.: "verde face laranja D opalino").

    Regra de ordem canonica:
      <cor-base> <D|DD> <face-laranja/pale-headed/slaty/violeta> <opalino> <mut-primaria> [misty DF]

    Regras especiais "violet":
      - Roseicollis: standalone "violet" (sem outra cor-base) => "*blue* D violeta"
      - Eyering:     standalone "violet" (sem outra cor-base) => "violeta" (sem D)
      - "violet factored [D] aqua"        => "aqua D violeta"  (ambos)
      - "violet turquoise" (so Roseicollis)=> "turquesa D violeta"
      - "violet Blue1Blue2" (so Eyering)  => "Azul1Azul2 violeta" (sem D)
    """
    t = texto_en.strip()

    # 0) Preprocessamento das variantes "violet" -> tokens atomicos
    #    para nao serem partidos pela reordenacao mais abaixo.

    # violet factored (D) aqua -> aqua D violeta
    t = re.sub(r"violet factored\s+(?:D\s+)?aqua", "__AQUA_DVIOLETA__", t)

    # violet turquoise -> turquesa D violeta  (so Roseicollis usa turquoise)
    t = re.sub(r"violet turquoise", "__TURQUESA_DVIOLETA__", t)

    # violet Blue1Blue2 -> Azul1Azul2 violeta  (Eyering, sem D)
    t = re.sub(r"violet Blue1Blue2", "__AZUL1AZUL2_VIOLETA__", t)

    # violet standalone: se nao ha outra cor-base explicita no texto
    tem_cor_base = re.search(r"\b(green|blue|aqua|turquoise|Blue1Blue2|teal)\b", t)
    if not tem_cor_base and re.search(r"\bviolet\b", t):
        if especie == "roseicollis":
            t = re.sub(r"\bviolet\b", "__BLUE_DVIOLETA__", t)
        else:
            # Eyering: violeta simples
            t = re.sub(r"\bviolet\b", "__VIOLETA_ONLY__", t)

    # "DM jade" -> "DF jade" (mas guardar como "jade DF")
    t = re.sub(r"DM jade", "__JADE_DF__", t)

    # tokens especiais mantidos
    special_tokens = {
        "SL ino": "SL ino",
        "NSL ino": "NSL ino",
        "SL SF greywing": "__SL_SF_GW__",
        "SL DF greywing": "__SL_DF_GW__",
        "SF edged": "__SF_EDGED__",
        "DF edged": "__DF_EDGED__",
        "SF euwing": "__SF_EUWING__",
        "DF euwing": "__DF_EUWING__",
        "DF misty": "__DF_MISTY__",
        "opaline-ino": "__OP_INO__",
        "cinnamon-ino": "__CIN_INO__",
        "opaline-cinnamon": "__OP_CIN__",
        "opaline-pallid": "__OP_PALLID__",
        "opaline-pale": "__OP_PALE__",
        "dominant pied": "__DOM_PIED__",
        "recessive pied": "__REC_PIED__",
        "orange face": "__ORANGE_FACE__",
        "pale headed": "__PALE_HEADED__",
        "bronze fallow": "__BRONZE_FALLOW__",
        "pale fallow": "__PALE_FALLOW__",
        "dun fallow": "__DUN_FALLOW__",
    }
    for k, v in special_tokens.items():
        t = _substitute_word(t, k, v)

    # cores base
    t = _substitute_word(t, "green", "verde")
    t = _substitute_word(t, "turquoise", "turquesa")
    t = _substitute_word(t, "aqua", "aqua")
    t = _substitute_word(t, "teal", "teal")
    # blue (depende de especie)
    t = _substitute_word(t, "blue", blue_pt(especie))
    # Blue1Blue2 (so eyering)
    t = _substitute_word(t, "Blue1Blue2", "Azul1Azul2")

    # modificadores simples
    t = _substitute_word(t, "opaline", "opalino")
    t = _substitute_word(t, "violet", "violeta")
    t = _substitute_word(t, "slaty", "slaty")
    t = _substitute_word(t, "cinnamon", "canela")
    t = _substitute_word(t, "pallid", "pálido")

    # mutacoes primarias
    t = _substitute_word(t, "dilute", "diluido")
    t = _substitute_word(t, "pastel", "pastel")
    t = _substitute_word(t, "crested", "crested")
    t = _substitute_word(t, "marbled", "marbled")
    t = _substitute_word(t, "misty", "misty")
    t = _substitute_word(t, "dec", "dec")

    # restaurar tokens especiais
    reverse = {
        "__AQUA_DVIOLETA__": "aqua D violeta",
        "__TURQUESA_DVIOLETA__": "turquesa D violeta",
        "__AZUL1AZUL2_VIOLETA__": "Azul1Azul2 violeta",
        "__BLUE_DVIOLETA__": "*blue* D violeta",
        "__VIOLETA_ONLY__": "violeta",
        "__JADE_DF__": "jade DF",
        "__SL_SF_GW__": "SL greywing SF",
        "__SL_DF_GW__": "SL greywing DF",
        "__SF_EDGED__": "edged SF",
        "__DF_EDGED__": "edged DF",
        "__SF_EUWING__": "euwing SF",
        "__DF_EUWING__": "euwing DF",
        "__DF_MISTY__": "misty DF",
        "__OP_INO__": "opalino-ino",
        "__CIN_INO__": "canela-ino",
        "__OP_CIN__": "opalino-canela",
        "__OP_PALLID__": "opalino-pálido",
        "__OP_PALE__": "opalino-pale",
        "__DOM_PIED__": "arlequim dominante",
        "__REC_PIED__": "arlequim recessivo",
        "__ORANGE_FACE__": "face laranja",
        "__PALE_HEADED__": "pale headed",
        "__BRONZE_FALLOW__": "bronze fallow",
        "__PALE_FALLOW__": "pale fallow",
        "__DUN_FALLOW__": "dun fallow",
    }
    # sexo
    t = _substitute_word(t, "male", "macho")
    t = _substitute_word(t, "female", "fêmea")

    # reordenacao canonica: split -> sort por categoria -> join
    # Todos os multi-word ficam como placeholders atomicos ate ao sort;
    # (SL ino / NSL ino ficam como 2 tokens da mesma categoria => ordem preservada).
    tokens = t.split()
    tokens_ordenados = _sort_tokens_canonical(tokens)
    t = " ".join(tokens_ordenados)

    # expandir placeholders atomicos ao final
    for placeholder, expansao in reverse.items():
        t = t.replace(placeholder, expansao)

    # normalizar espacos
    t = re.sub(r"\s+", " ", t).strip()
    return t


def traduzir_conjunto(texto_en: str, especie: str) -> str:
    """
    Uma entrada de codigo pode ter varias sub-entradas separadas por virgula.
    Ex.: "orange face green, orange face D green, orange face DD green"
    Traduz cada uma e junta com ', '.

    Se a especie exige prefixo de sexo (Taranta/Canus/Pullarius: "male"/"female"),
    o sexo so aparece na primeira sub-entrada (estilo 2024).
    """
    partes = [p.strip() for p in texto_en.split(",") if p.strip()]
    traduzidas = [traduzir_mutacao(p, especie) for p in partes]
    # Dedupe do sexo: se todas as partes comecam com "macho" ou "fêmea",
    # mantem so a primeira e remove das seguintes.
    for sexo in ("macho", "fêmea"):
        if all(p.startswith(sexo + " ") for p in traduzidas) and len(traduzidas) > 1:
            traduzidas = [traduzidas[0]] + [p[len(sexo) + 1:] for p in traduzidas[1:]]
            break
    return ", ".join(traduzidas)


def traduzir_serie(nome_en: str, especie: str = "") -> str:
    """Traduz o nome de uma serie/grupo em teams equipas (birds in <serie>)."""
    m = nome_en.strip().lower()
    mapa = {
        "greenseries": "verde",
        "tealseries": "teal",
        "greenserie": "verde",
        "dominant edged": "edged dominante",
        "dominant pied": "arlequim dominante",
        "recessive pied": "arlequim recessivo",
        "rare mutations": "mutações raras",
        "sl dominant greywing": "SL greywing dominante",
        "nsl ino": "NSL ino",
        "sl ino": "SL ino",
        "bronze fallow": "bronze fallow",
        "pale fallow": "pale fallow",
        "dun fallow": "dun fallow",
        "dilute": "diluido",
        "misty": "misty",
        "euwing": "euwing",
        "pastel": "pastel",
        "dec": "dec",
        "pale": "pale",
        "crested": "crested",
        "marbled": "marbled",
        "cinnamon": "canela",
        "pallid": "pálido",
        "dm jade": "jade DF",
        "aqua": "aqua",
        "blue": blue_pt(especie) if especie else "azul",
        "turquoise": "turquesa",
        "blue1blue2": "Azul1Azul2",
    }
    return mapa.get(m, m)


def traduzir_crested_combinacao(desc_en: str) -> str:
    """
    Traduz 'crested in combination with (the) birds in groups X-Y; ...'
    e variantes 'opaline crested', 'DF dominant reduced', etc.
    Devolve o texto ja em PT, para colar depois de "A. Especie".
    """
    t = desc_en.strip()
    # normalizacoes de sequencia inteira
    t = re.sub(r"in combination with (the )?birds in group\s+(\d+)\s+till\s+group\s+(\d+)",
               r"em combinação com aves dos grupos \2 ao \3", t, flags=re.IGNORECASE)
    t = re.sub(r"in combination with (the )?birds in groups?", "em combinação com aves dos grupos", t, flags=re.IGNORECASE)
    t = re.sub(r"in combination with birds", "em combinação com aves", t, flags=re.IGNORECASE)
    # DF dominant reduced -> DF reduced dominante
    t = re.sub(r"opaline\s+DF\s+dominant\s+reduced", "DF reduced dominante opalino", t, flags=re.IGNORECASE)
    t = re.sub(r"DF\s+dominant\s+reduced", "DF reduced dominante", t, flags=re.IGNORECASE)
    # opaline crested
    t = re.sub(r"^opaline\s+crested\b", "crested opalino", t, flags=re.IGNORECASE)
    t = re.sub(r"\bopaline\s+crested\b", "crested opalino", t, flags=re.IGNORECASE)
    # crested standalone: manter
    return t


def traduzir_titulo_grupo(titulo_en: str, especie: str) -> str:
    """
    Ex.: "GROUP AGAPORNIS ROSEICOLLIS GREENSERIES"
      -> "GRUPO AGAPORNIS ROSEICOLLIS SÉRIE VERDE"
    """
    t = titulo_en.upper()
    t = t.replace("GROUP ", "GRUPO ")
    t = t.replace("GREENSERIES", "SÉRIE VERDE")
    t = t.replace("TEALSERIES", "SÉRIE TEAL")
    t = t.replace("GREEN", "VERDE")
    t = t.replace("BLUE1BLUE2", "AZUL1AZUL2")
    # blue: depende de especie
    if especie == "roseicollis":
        t = re.sub(r"\bBLUE\b", "*BLUE*", t)
    else:
        t = re.sub(r"\bBLUE\b", "AZUL", t)
    t = t.replace("TURQUOISE", "TURQUESA")
    t = t.replace("MARBLED", "MARBLED")
    t = t.replace("DILUTE", "DILUIDO")
    t = t.replace("CINNAMON", "CANELA")
    t = t.replace("PALLID", "PÁLIDO")
    t = t.replace("DOMINANT PIED", "ARLEQUIM DOMINANTE")
    t = t.replace("RECESSIVE PIED", "ARLEQUIM RECESSIVO")
    t = t.replace("DM JADE", "JADE DF")
    t = t.replace("DOMINANT EDGED", "EDGED DOMINANTE")
    t = t.replace("SL DOMINANT GREYWING", "SL GREYWING DOMINANTE")
    t = t.replace("RARE MUTATIONS", "MUTAÇÕES RARAS")
    t = t.replace("MUTATIONS", "MUTAÇÕES")
    # normalizar separador "A. FISCHERI"
    t = re.sub(r"\s+-\s+", " – ", t)
    t = re.sub(r"[��]", "–", t)  # substitui replacement char
    return t


# ============================================================
# PARSER DO TXT
# ============================================================

# blocos: cada grupo comeca com "GROUP ..." ou "GROUP AGAPORNIS PERSONATUS - ..."
# depois vem cabecalhos com "A Color" ou "A A A A Color" e depois linhas de codigo.
# tambem existem seccoes STUDY GROUP, TEAMS, e notas finais.

RE_GROUP = re.compile(r"^\s*GROUP\s+(.+?)\s*$")
RE_STUDY = re.compile(r"^\s*STUDY GROUP", re.IGNORECASE)
RE_TEAMS = re.compile(r"^\s*TEAMS,\s*T CLASS\s+(.+?)\s*$", re.IGNORECASE)
# codigos:
#   simples:   "001/01 green"
#   4-esp:     "050/01  100/01         150/01        200/01 green"
#   mistos com NYA
RE_CODIGO = re.compile(
    r"^\s*(?:(\d{3}/\d{2})|NYA)\s+"
    r"(?:(?:(\d{3}/\d{2})|NYA)\s+)?"
    r"(?:(?:(\d{3}/\d{2})|NYA)\s+)?"
    r"(?:(?:(\d{3}/\d{2})|NYA)\s+)?"
    r"(.*)$"
)


def identifica_especie_do_grupo(titulo_en: str) -> list[str]:
    """Devolve lista de especies (ordenada) a que o grupo se aplica."""
    t = titulo_en.upper()
    especies = []
    if "ROSEICOLLIS" in t:
        especies.append("roseicollis")
    if "PERSONATUS" in t or "A. FISCHERI" in t or "A. NIGRIGENIS" in t or "A. LILIANAE" in t:
        # grupo eyering multi-especie
        for e in ("personatus", "fischeri", "nigrigenis", "lilianae"):
            if e not in especies:
                especies.append(e)
        # remove roseicollis (nao aplicavel a grupos eyering)
        especies = [e for e in especies if e != "roseicollis" or "ROSEICOLLIS" in t]
    if "CANUS" in t and "canus" not in especies:
        especies.append("canus")
    if "TARANTA" in t and "taranta" not in especies:
        especies.append("taranta")
    if "PULLARIUS" in t and "pullarius" not in especies:
        especies.append("pullarius")
    return especies


def parse_txt(path: Path):
    """
    Parseia o txt e devolve uma lista de blocos:
      [ {tipo: 'group'|'study'|'teams', titulo_en, especies, entries: [ (codigos_dict, texto_en) ] }, ... ]
    codigos_dict: {especie: "xxx/yy" or None}
    """
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()

    blocks = []
    current = None
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if not stripped:
            i += 1
            continue

        # Detecta cabecalhos de bloco
        if stripped.upper().startswith("GROUP "):
            titulo = stripped[len("GROUP "):].strip()
            especies = identifica_especie_do_grupo(titulo)
            current = {"tipo": "group", "titulo_en": titulo, "especies": especies, "entries": []}
            blocks.append(current)
            i += 1
            continue

        if RE_STUDY.match(stripped):
            titulo = stripped
            current = {"tipo": "study", "titulo_en": titulo, "especies": [], "entries": []}
            blocks.append(current)
            i += 1
            continue

        m_teams = RE_TEAMS.match(stripped)
        if m_teams:
            titulo = m_teams.group(1).strip()
            especies = identifica_especie_do_grupo(titulo)
            current = {"tipo": "teams", "titulo_en": titulo, "especies": especies, "entries": []}
            blocks.append(current)
            i += 1
            continue

        # ignora linhas de cabecalho de tabela ("A Color", "A. personatus A. fischeri ...")
        if re.fullmatch(r"[A|N|T]\s+Color", stripped) or "Color" in stripped and len(stripped) < 80:
            i += 1
            continue
        if re.match(r"^A\.\s*(personatus|fischeri|nigrigenis|lilianae|Canus|Taranta|Pullarius)", stripped, re.IGNORECASE):
            i += 1
            continue
        if re.fullmatch(r"[ATN\s]+", stripped):  # linhas so com A A A A ou T T T T
            i += 1
            continue
        if stripped.startswith("Canus Taranta"):
            i += 1
            continue

        # tenta parsear linha de codigo
        if current is not None:
            # Pode ser linha multi-especie com <=4 codigos, ou single-especie com 1 codigo
            # Estrategia: extrair todos os tokens que sao "xxx/yy" ou "NYA" no inicio
            m = re.match(r"^([A-Za-z]*)\s*", line)
            # extrai tokens do inicio
            rest = stripped
            codigos = []
            while True:
                m = re.match(r"^\s*(\d{3}/\d{2}|NYA)\b\s*", rest)
                if not m:
                    break
                codigos.append(m.group(1))
                rest = rest[m.end():]
            if codigos:
                descricao = rest.strip()
                # se descricao vazia, pode ser continuacao na linha seguinte
                if not descricao and i + 1 < len(lines):
                    nxt = lines[i + 1].strip()
                    if nxt and not nxt.upper().startswith("GROUP") and not RE_TEAMS.match(nxt):
                        # nao mistura com codigo novo
                        if not re.match(r"^\s*(\d{3}/\d{2}|NYA)\b", nxt):
                            descricao = nxt
                            i += 1
                # atribui codigos as especies
                if current["tipo"] == "group" and len(current["especies"]) == 1:
                    esp = current["especies"][0]
                    codigos_dict = {esp: codigos[0] if codigos[0] != "NYA" else None}
                elif current["tipo"] == "group" and len(current["especies"]) >= 2:
                    # eyering multi: ordem personatus, fischeri, nigrigenis, lilianae
                    order = ["personatus", "fischeri", "nigrigenis", "lilianae"]
                    codigos_dict = {}
                    for idx, esp in enumerate(order):
                        c = codigos[idx] if idx < len(codigos) else None
                        codigos_dict[esp] = c if c and c != "NYA" else None
                elif current["tipo"] == "teams":
                    # tenta atribuir consoante ordem das especies do titulo
                    codigos_dict = {}
                    for idx, esp in enumerate(current["especies"]):
                        c = codigos[idx] if idx < len(codigos) else None
                        codigos_dict[esp] = c if c and c != "NYA" else None
                elif current["tipo"] == "study":
                    codigos_dict = {"study": codigos[0] if codigos[0] != "NYA" else None}
                else:
                    codigos_dict = {}
                current["entries"].append((codigos_dict, descricao))
                i += 1
                continue

        i += 1
    return blocks


# ============================================================
# GERADOR DO XLSX
# ============================================================

def prepara_nome_grupo_pt(bloco) -> str:
    """Titulo do grupo em PT + " - INDIVIDUAL" / " - EQUIPAS (4 AVES)"."""
    tit = traduzir_titulo_grupo(bloco["titulo_en"],
                                 bloco["especies"][0] if bloco["especies"] else "roseicollis")
    if bloco["tipo"] == "group":
        return f"{tit} - INDIVIDUAL"
    if bloco["tipo"] == "teams":
        return f"{tit} - EQUIPAS (4 AVES)"
    if bloco["tipo"] == "study":
        return "GRUPO DE ESTUDO (Todas as novas mutações. Estas aves não competem para 'Best in Show') - INDIVIDUAL"
    return tit


def gerar_xlsx(blocks, out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folha1"

    # estilos
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    fill_title = PatternFill("solid", fgColor="1F4E78")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    font_group = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_group = PatternFill("solid", fgColor="2E75B6")

    font_hdr = Font(name="Calibri", size=10, bold=True)
    fill_hdr = PatternFill("solid", fgColor="DDEBF7")

    font_body = Font(name="Calibri", size=10)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center_body = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 90

    row = 1
    # Cabecalho principal
    ws.cell(row, 1, "9ª Exposição BVA PORTUGAL / AOB 2026 de Agapornis & Forpus")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 1).font = font_title
    ws.cell(row, 1).fill = fill_title
    ws.cell(row, 1).alignment = align_center
    ws.row_dimensions[row].height = 26
    row += 2

    # numeracao unica global monotonica.
    # proximo_seccao(par=True) devolve o proximo numero par >= cursor, e avanca.
    # proximo_seccao(par=False) devolve o proximo numero impar >= cursor, e avanca.
    _cursor = [1]  # lista para mutabilidade em closure

    def proximo_seccao(par: bool) -> int:
        n = _cursor[0]
        if par:
            if n % 2 != 0:
                n += 1
        else:
            if n % 2 == 0:
                n += 1
        _cursor[0] = n + 1
        return n

    def escrever_group_header(titulo_pt: str):
        nonlocal row
        ws.cell(row, 1, titulo_pt)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1).font = font_group
        ws.cell(row, 1).fill = fill_group
        ws.cell(row, 1).alignment = align_center
        ws.row_dimensions[row].height = 22
        row += 1
        # cabecalho da tabela
        for col, txt in enumerate(["Classe", "Secção AOB", "Código BVA", "Mutação"], 1):
            c = ws.cell(row, col, txt)
            c.font = font_hdr
            c.fill = fill_hdr
            c.alignment = align_center
            c.border = border
        row += 1

    def escrever_linha(classe: str, seccao: int, codigo: str, texto: str):
        nonlocal row
        ws.cell(row, 1, classe).font = font_body
        ws.cell(row, 1).alignment = align_center_body
        ws.cell(row, 1).border = border
        ws.cell(row, 2, seccao).font = font_body
        ws.cell(row, 2).alignment = align_center_body
        ws.cell(row, 2).border = border
        ws.cell(row, 3, codigo or "").font = font_body
        ws.cell(row, 3).alignment = align_center_body
        ws.cell(row, 3).border = border
        ws.cell(row, 4, texto).font = font_body
        ws.cell(row, 4).alignment = align_left
        ws.cell(row, 4).border = border
        row += 1

    # ---------------------------------------------------------
    # REORDENACAO POR ESPECIE (em vez de por bloco/mutacao)
    # ---------------------------------------------------------
    # Estrutura pretendida (igual ao Excel 2024):
    #   Para cada especie na ordem: roseicollis, personatus, fischeri,
    #   nigrigenis, lilianae, canus, taranta, pullarius:
    #     - todos os grupos INDIVIDUAIS (numeros pares)
    #     - todas as equipas dessa especie      (numeros impares apos os pares)
    #   No fim: STUDY GROUP (classe N).
    #
    # Numeracao:
    #   - Individuais: proximo par (contador_par)
    #   - Equipas: apos o ultimo par da especie, saltar para o proximo impar
    SPECIES_ORDER = {
        "roseicollis": 1, "personatus": 2, "fischeri": 3, "nigrigenis": 4,
        "lilianae": 5, "canus": 6, "taranta": 7, "pullarius": 8,
    }
    # Constroi lista plana: [(especie_prio, tipo_prio, bidx, bloco, especie)]
    #   tipo_prio: 0=group, 1=teams (individuais antes de equipas por especie)
    planos = []
    for bidx, bloco in enumerate(blocks):
        if bloco["tipo"] == "group":
            for especie in bloco["especies"]:
                prio = SPECIES_ORDER.get(especie, 99)
                planos.append((prio, 0, bidx, bloco, especie))
        elif bloco["tipo"] == "teams":
            for especie in bloco["especies"]:
                prio = SPECIES_ORDER.get(especie, 99)
                planos.append((prio, 1, bidx, bloco, especie))
        elif bloco["tipo"] == "study":
            planos.append((100, 0, bidx, bloco, None))  # study no fim
    planos.sort(key=lambda x: (x[0], x[1], x[2]))

    for prio_esp, tipo_prio, bidx, bloco, especie in planos:
        if bloco["tipo"] in ("group", "teams"):
            # loop interno reduzido a uma unica iteracao (a especie ja veio decidida)
            for _especie_ in [especie]:
                especie = _especie_
                # verifica se pelo menos uma entry tem codigo para esta especie
                entradas_especie = []
                for codigos_dict, desc in bloco["entries"]:
                    codigo = codigos_dict.get(especie)
                    if codigo is None:
                        continue  # NYA -> omitir
                    if not desc or not desc.strip():
                        continue  # sem descricao -> omitir
                    entradas_especie.append((codigo, desc))
                if not entradas_especie:
                    continue

                # titulo do grupo (troca AGAPORNIS PERSONATUS - A. FISCHERI ... por so a especie)
                titulo_base = traduzir_titulo_grupo(bloco["titulo_en"], especie)
                # se e multi-especie, substitui todo o prefixo pelo AGAPORNIS <especie>
                # padroes: "AGAPORNIS PERSONATUS – A. FISCHERI – A. NIGRIGENIS – A. LILIANAE ..."
                #          "A. PERSONATUS – A. FISCHERI – A. NIGRIGENIS –A. LILIANAE ..."
                #          "A. CANUS – A. TARANTA – A. PULLARIUS ..."
                titulo_base = re.sub(
                    r"(?:AGAPORNIS\s+)?A\.\s*PERSONATUS.*?LILIANAE",
                    f"AGAPORNIS {especie.upper()}",
                    titulo_base,
                )
                titulo_base = re.sub(
                    r"AGAPORNIS\s+PERSONATUS.*?LILIANAE",
                    f"AGAPORNIS {especie.upper()}",
                    titulo_base,
                )
                titulo_base = re.sub(
                    r"(?:AGAPORNIS\s+)?A\.\s*CANUS.*?PULLARIUS",
                    f"AGAPORNIS {especie.upper()}",
                    titulo_base,
                )
                # limpeza: se comeca com "A. ROSEICOLLIS" (nas equipas), passa a "AGAPORNIS ROSEICOLLIS"
                titulo_base = re.sub(r"^A\.\s+ROSEICOLLIS", "AGAPORNIS ROSEICOLLIS", titulo_base)
                titulo_base = re.sub(r"^AGAPORNIS\s+TARANTA\s+MUTAÇÕES", "AGAPORNIS TARANTA MUTAÇÕES", titulo_base)
                # tambem para grupos so com uma especie manter tal como esta
                if bloco["tipo"] == "group":
                    titulo_pt = f"{titulo_base} - INDIVIDUAL"
                else:
                    # equipas
                    titulo_pt = f"{titulo_base} - EQUIPAS (4 AVES)"

                escrever_group_header(titulo_pt)

                # classe unica: J (todas as inscricoes AOB usam classe J)
                classe = "J"

                for codigo, desc in entradas_especie:
                    # regra especial: remover "aves da serie violeta" das equipas eyering
                    if bloco["tipo"] == "teams" and especie in ("personatus", "fischeri", "nigrigenis", "lilianae"):
                        if re.search(r"\bviolet\b", desc, re.IGNORECASE) and "birds in" in desc.lower():
                            continue

                    # tratamento especial para equipas
                    if bloco["tipo"] == "teams":
                        d = desc.strip()
                        dl = d.lower()
                        # Roseicollis: "A. roseicollis <serie> (group X)"
                        m_rose = re.match(r"^A\.\s*roseicollis\s+(.+?)\s*\(group\s+\d+\)\s*$", d, re.IGNORECASE)
                        if m_rose:
                            serie_en = m_rose.group(1).strip()
                            texto_pt = f"{ESPECIES_TITLE[especie]} aves da série {traduzir_serie(serie_en, especie)}"
                        elif dl.startswith("birds in"):
                            serie_en = d[len("birds in"):].strip()
                            texto_pt = f"{ESPECIES_TITLE[especie]} aves da série {traduzir_serie(serie_en, especie)}"
                        elif dl == "green":
                            texto_pt = f"{ESPECIES_TITLE[especie]} verde"
                        elif dl == "blue":
                            texto_pt = f"{ESPECIES_TITLE[especie]} {blue_pt(especie)}"
                        elif dl == "aqua":
                            texto_pt = f"{ESPECIES_TITLE[especie]} aqua"
                        elif dl == "blue1blue2":
                            texto_pt = f"{ESPECIES_TITLE[especie]} Azul1Azul2"
                        else:
                            texto_pt = f"{ESPECIES_TITLE[especie]} {traduzir_conjunto(d, especie)}"
                    else:
                        # grupos: verifica se e descricao especial (crested combinacao / rare mutations)
                        dl = desc.lower().strip()
                        if "in combination with" in dl:
                            texto_pt = f"{ESPECIES_TITLE[especie]} {traduzir_crested_combinacao(desc)}"
                        else:
                            texto_pt = f"{ESPECIES_TITLE[especie]} {traduzir_conjunto(desc, especie)}"

                    seccao = proximo_seccao(par=(bloco["tipo"] == "group"))
                    escrever_linha(classe, seccao, codigo, texto_pt)

        elif bloco["tipo"] == "study":
            escrever_group_header(prepara_nome_grupo_pt(bloco))
            for codigos_dict, desc in bloco["entries"]:
                cod = codigos_dict.get("study")
                if cod is None:
                    continue
                # desc = "A. roseicollis" -> "A. Roseicollis"
                nome_pt = desc.strip()
                for k, v in ESPECIES_TITLE.items():
                    nome_pt = re.sub(rf"A\.\s*{k}", v, nome_pt, flags=re.IGNORECASE)
                escrever_linha("J", proximo_seccao(par=True), cod, nome_pt)

    # ---- notas finais (traduzidas do PDF) ----
    row += 1
    notas = [
        "Notas:",
        "Violeta: um violeta é um azul portador de um factor escuro (D) e um ou dois factores violeta.",
        "Violeta Azul1Azul2 e violeta turquesa: são portadores de um factor escuro e um ou dois factores violeta.",
        "Aqua D violeta (violet factored aqua): aqua combinado com um factor escuro e um ou dois factores violeta.",
        "Misty: aceite exclusivamente em verde, azul, turquesa ou aqua e tem de possuir dois factores misty. A combinação misty + factor escuro NÃO é aceite como forma de exposição.",
        "Slaty: tem de possuir no máximo um factor escuro e um ou dois factores slaty. Nenhuma outra combinação é aceite como forma de exposição.",
        "Dec e pastel: apenas homozigotos são aceites.",
        "Azul1Azul2: turquesa fenotípico nas espécies com anel ocular (eyering).",
        "NYA: Not Yet Accepted (Ainda Não Aceite). Combinações NYA foram omitidas desta lista.",
    ]
    for nota in notas:
        ws.cell(row, 1, nota).font = Font(italic=True, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    wb.save(out_path)
    print(f"Escrito: {out_path}")
    print(f"Total linhas: {row}")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    blocks = parse_txt(TXT)
    print(f"Blocos parseados: {len(blocks)}")
    for b in blocks[:5]:
        print(f"  - [{b['tipo']}] {b['titulo_en'][:70]}  especies={b['especies']}  n_entradas={len(b['entries'])}")
    total_entries = sum(len(b["entries"]) for b in blocks)
    print(f"Total entradas: {total_entries}")
    gerar_xlsx(blocks, OUT)
