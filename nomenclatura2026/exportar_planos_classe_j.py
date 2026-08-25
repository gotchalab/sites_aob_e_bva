# -*- coding: utf-8 -*-
"""
Le o Classes_Agapornis_2026_PT.xlsx (gerado por gerar_nomenclatura_2026_pt.py)
e produz dois ficheiros planos:
  - Classes_Agapornis_2026_INDIVIDUAIS_J.xlsx  (classe A -> J)
  - Classes_Agapornis_2026_EQUIPAS_J.xlsx       (classe T -> J)

Ambos:
  - Sem cabecalhos de grupo repetidos no meio das linhas
  - Sem linhas de header repetidas por grupo ("Classe | Seccao AOB | ...")
  - Coluna Classe = "J" em todas as linhas
  - Colunas: Classe | Seccao AOB | Grupo | Mutacao | Codigo BVA
  - Uma unica linha de header no topo
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).parent
SRC = BASE / "Classes_Agapornis_2026_PT.xlsx"
OUT_INDIV = BASE / "Classes_Agapornis_2026_INDIVIDUAIS_J.xlsx"
OUT_EQUIP = BASE / "Classes_Agapornis_2026_EQUIPAS_J.xlsx"


def normalizar_grupo(header: str) -> str:
    """Extrai o nome do grupo do header (ex.: "AGAPORNIS ROSEICOLLIS VERDE - INDIVIDUAL"
    -> "Roseicollis Verde"). Preserva formato leitura amigavel."""
    t = header.strip()
    # caso especial: grupo de estudo
    if t.upper().startswith("GRUPO DE ESTUDO"):
        return "Grupo de Estudo"
    # remove sufixos
    for suf in (" - INDIVIDUAL", " - EQUIPAS (4 AVES)"):
        if t.endswith(suf):
            t = t[: -len(suf)]
            break
    # tira "AGAPORNIS " prefixo
    if t.upper().startswith("AGAPORNIS "):
        t = t[len("AGAPORNIS "):]
    # capitalize palavra-a-palavra, mas mantendo tokens especiais e stopwords minusculas
    partes = t.split()
    especiais = {"SF", "DF", "SL", "NSL", "DM", "DEC", "NYA"}
    stopwords_pt = {"de", "da", "do", "das", "dos", "e", "em", "para", "com"}
    resultado = []
    for i, p in enumerate(partes):
        if p.upper() in especiais:
            resultado.append(p.upper())
        elif p.startswith("*"):
            resultado.append(p.lower())
        elif i > 0 and p.lower() in stopwords_pt:
            resultado.append(p.lower())
        else:
            resultado.append(p.capitalize())
    return " ".join(resultado)


def extrair_por_tipo(src_path: Path, tipo: str) -> list[dict]:
    """Le o Excel fonte e devolve linhas cujo grupo actual e do tipo pedido:
       tipo = 'INDIVIDUAL' | 'EQUIPAS' | 'ESTUDO'"""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb["Folha1"]
    linhas = []
    grupo_actual = ""
    tipo_actual = ""
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        # detecta cabecalho de grupo
        if a and isinstance(a, str):
            up = a.upper()
            if up.startswith("AGAPORNIS "):
                grupo_actual = normalizar_grupo(a)
                if "EQUIPAS" in up:
                    tipo_actual = "EQUIPAS"
                else:
                    tipo_actual = "INDIVIDUAL"
                continue
            if up.startswith("GRUPO DE ESTUDO"):
                grupo_actual = normalizar_grupo(a)
                tipo_actual = "ESTUDO"
                continue
        # dados: linhas com classe (J) e seccao numerica
        if a == "J" and isinstance(b, (int, float)) and tipo_actual == tipo:
            linhas.append({
                "seccao": b,
                "codigo": c or "",
                "mutacao": d or "",
                "grupo": grupo_actual,
            })
    return linhas


def gerar_xlsx_plano(linhas: list[dict], out_path: Path, titulo: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folha1"

    # estilos
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    fill_title = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(name="Calibri", size=10, bold=True)
    fill_hdr = PatternFill("solid", fgColor="DDEBF7")
    font_body = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 85
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35

    # titulo
    ws.cell(1, 1, titulo)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(1, 1).font = font_title
    ws.cell(1, 1).fill = fill_title
    ws.cell(1, 1).alignment = align_center
    ws.row_dimensions[1].height = 26

    # cabecalho
    for col, txt in enumerate(["Classe", "Secção AOB", "Mutação", "Código BVA", "Grupo"], 1):
        c = ws.cell(3, col, txt)
        c.font = font_hdr
        c.fill = fill_hdr
        c.alignment = align_center
        c.border = border

    # linhas
    row = 4
    for linha in linhas:
        cells = [
            ("J", align_center),
            (linha["seccao"], align_center),
            (linha["mutacao"], align_left),
            (linha["codigo"], align_center),
            (linha["grupo"], align_left),
        ]
        for col, (val, align) in enumerate(cells, 1):
            c = ws.cell(row, col, val)
            c.font = font_body
            c.alignment = align
            c.border = border
        row += 1

    wb.save(out_path)
    print(f"Escrito: {out_path}  ({len(linhas)} linhas)")


if __name__ == "__main__":
    if not SRC.exists():
        raise SystemExit(f"Ficheiro fonte nao encontrado: {SRC}")

    indivs = extrair_por_tipo(SRC, "INDIVIDUAL")
    equipas = extrair_por_tipo(SRC, "EQUIPAS")
    # Study group tambem vai no ficheiro individual
    study = extrair_por_tipo(SRC, "ESTUDO")
    indivs.extend(study)

    gerar_xlsx_plano(
        indivs,
        OUT_INDIV,
        "9ª Exposição BVA PORTUGAL / AOB 2026 — INDIVIDUAIS (Classe J)",
    )
    gerar_xlsx_plano(
        equipas,
        OUT_EQUIP,
        "9ª Exposição BVA PORTUGAL / AOB 2026 — EQUIPAS 4 AVES (Classe J)",
    )
